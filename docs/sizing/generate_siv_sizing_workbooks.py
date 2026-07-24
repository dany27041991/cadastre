#!/usr/bin/env python3
"""
Generate SIV sizing workbooks by cloning SIM templates and inserting SIV rows only.

Approach A: clone Fabbisogno_SIM_Globale.xlsx and dataiku_rsc_cu.xlsx preserving
structure (Foglio1, AutoFilter, Excel Tables, styles) and replace data rows with SIV entries.
Foglio1 is left unchanged from the SIM template (lookup lists, layout).
"""

from __future__ import annotations

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[3]
OUT_DIR = ROOT / "cadastre" / "docs" / "sizing"
DESKTOP = Path.home() / "Desktop"
ESEMPI_DIR = Path(
    "/Users/danilogiovannico/Desktop/LAVORO/MASE/SIV/Documentazione/DIMENSIONAMENTO/ESEMPI"
)

TEMPLATE_FABBISOGNO = ESEMPI_DIR / "Fabbisogno_SIM_Globale.xlsx"
TEMPLATE_FABBISOGNO_FALLBACKS = (
    DESKTOP / "Fabbisogno_SIM_Globale.xlsx",
    OUT_DIR / "Fabbisogno_SIV_Catasto_Arboreo.xlsx",
)
TEMPLATE_DATAIKU = DESKTOP / "dataiku_rsc_cu.xlsx"

OUT_FABBISOGNO = OUT_DIR / "Fabbisogno_SIV_Catasto_Arboreo.xlsx"
OUT_DATAIKU = OUT_DIR / "dataiku_rsc_siv_catasto.xlsx"

SOC = "DXC"
CU = "SIV"
DATA_COLL = datetime(2026, 6, 30)
DATA_PROD = datetime(2026, 9, 30)

# Container: PVC locale per replica (log, temp, export staging, headroom futuro).
# Dati persistenti di dominio restano su PostGIS e MinIO.
SIV_RESOURCES: list[dict] = [
    {
        "servizio": "Container",
        "desc": "SIV - Quota cluster OC DXAP (Backend FastAPI)",
        "os": "Debian/Linux",
        "sw": None,
        "note": "HPA min 6 max 12 pod. PVC locale: log/temp/export staging.",
        "envs": {
            "Sviluppo": {"inc": "SI", "qty": 1, "cpu": 2, "ram": 4, "disk": 10, "ril": "No", "date": DATA_COLL},
            "Collaudo": {"inc": "SI", "qty": 2, "cpu": 2, "ram": 4, "disk": 20, "ril": "No", "date": DATA_COLL},
            "Produzione": {"inc": "SI", "qty": 6, "cpu": 4, "ram": 8, "disk": 50, "ril": "No", "date": DATA_PROD},
        },
    },
    {
        "servizio": "Container",
        "desc": "SIV - Quota cluster OC DXAP (Microfrontend React)",
        "os": "Debian/Linux",
        "sw": None,
        "note": "nginx + bundle statico; PVC locale per log/cache asset.",
        "envs": {
            "Sviluppo": {"inc": "SI", "qty": 1, "cpu": 1, "ram": 2, "disk": 5, "ril": "No", "date": DATA_COLL},
            "Collaudo": {"inc": "SI", "qty": 2, "cpu": 1, "ram": 2, "disk": 10, "ril": "No", "date": DATA_COLL},
            "Produzione": {"inc": "SI", "qty": 3, "cpu": 1, "ram": 2, "disk": 20, "ril": "No", "date": DATA_PROD},
        },
    },
    {
        "servizio": "Servizi PaaS DB",
        "desc": "SIV - PostgreSQL + PostGIS",
        "os": "Postgres release 16.x",
        "sw": "estensione postGIS",
        "note": "Partizionamento ISTAT; ~36M asset nazionali, replica RO in produzione.",
        "envs": {
            "Sviluppo": {"inc": "SI", "qty": 1, "cpu": 4, "ram": 8, "disk": 10, "ril": "No", "date": DATA_COLL},
            "Collaudo": {"inc": "SI", "qty": 1, "cpu": 8, "ram": 16, "disk": 100, "ril": "No", "date": DATA_COLL},
            "Produzione": {"inc": "SI", "qty": 1, "cpu": 16, "ram": 64, "disk": 500, "ril": "No", "date": DATA_PROD},
        },
    },
    {
        "servizio": "Storage MinIO",
        "desc": "SIV - Storage per MinIO",
        "os": "PaaS MinIO",
        "sw": None,
        "note": None,
        "envs": {
            "Sviluppo": {"inc": "SI", "qty": 1, "cpu": "NA", "ram": "NA", "disk": 10, "ril": "No", "date": DATA_COLL},
            "Collaudo": {"inc": "SI", "qty": 1, "cpu": "NA", "ram": "NA", "disk": 100, "ril": "No", "date": DATA_COLL},
            "Produzione": {"inc": "SI", "qty": 1, "cpu": "NA", "ram": "NA", "disk": 500, "ril": "No", "date": DATA_PROD},
        },
    },
]


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 16) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def _build_siv_rows() -> list[list]:
    rows: list[list] = []
    for res in SIV_RESOURCES:
        for env in ("Sviluppo", "Collaudo", "Produzione"):
            if env not in res["envs"]:
                continue
            cfg = res["envs"][env]
            rows.append(
                [
                    env,
                    SOC,
                    res["servizio"],
                    CU,
                    res["desc"],
                    cfg["inc"],
                    cfg["qty"],
                    res["os"],
                    cfg["cpu"],
                    cfg["ram"],
                    cfg["disk"],
                    cfg["date"],
                    cfg["ril"],
                    res.get("sw"),
                    None,
                    res.get("note") or "\xa0",
                ]
            )
    return rows


def _clear_data_rows(ws, first_data_row: int = 9) -> None:
    if ws.max_row >= first_data_row:
        ws.delete_rows(first_data_row, ws.max_row - first_data_row + 1)


def _write_data_rows(ws, rows: list[list], style_src_row: int = 9) -> None:
    start = 9
    for i, values in enumerate(rows):
        r = start + i
        if r <= ws.max_row:
            _copy_row_style(ws, style_src_row, r)
        for col, val in enumerate(values, start=1):
            ws.cell(r, col, val)
        ws.row_dimensions[r].hidden = False


def _reset_auto_filter(ws, last_data_row: int) -> None:
    """Remove template filter criteria (e.g. hidden Container rows) and reapply range."""
    ws.auto_filter.ref = None
    ws.auto_filter.filterColumn = []
    ws.auto_filter.ref = f"A8:P{last_data_row}"


def _update_sheet_post_write(ws, last_data_row: int) -> None:
    _reset_auto_filter(ws, last_data_row)
    if "Tabella1" in ws.tables:
        ws.tables["Tabella1"].ref = f"A6:P{last_data_row}"


def _add_rilasciato_validation(ws, last_data_row: int) -> None:
    """Dropdown col. M (Rilasciato) → Foglio1!B23:B24 (Si/No), come template SIM."""
    dv = DataValidation(
        type="list",
        formula1="='Foglio1'!$B$23:$B$24",
        allow_blank=True,
    )
    dv.add(f"M9:M{last_data_row}")
    ws.add_data_validation(dv)


def _resolve_fabbisogno_template() -> Path:
    if TEMPLATE_FABBISOGNO.exists():
        return TEMPLATE_FABBISOGNO
    for candidate in TEMPLATE_FABBISOGNO_FALLBACKS:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        f"Fabbisogno template not found: {TEMPLATE_FABBISOGNO} or fallbacks"
    )


def generate_fabbisogno() -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(_resolve_fabbisogno_template(), OUT_FABBISOGNO)

    wb = load_workbook(OUT_FABBISOGNO)

    rows = _build_siv_rows()
    last_row = 8 + len(rows)

    ws_c = wb["Fabbisogno Corrente"]
    style_row = 9 if ws_c.max_row >= 9 else 6
    _clear_data_rows(ws_c)
    _write_data_rows(ws_c, rows, style_row)
    _update_sheet_post_write(ws_c, last_row)
    _add_rilasciato_validation(ws_c, last_row)

    ws_f = wb["Fabbisogno Fine Progetto"]
    style_row_f = 9 if ws_f.max_row >= 9 else 6
    _clear_data_rows(ws_f)
    _write_data_rows(ws_f, rows, style_row_f)
    _update_sheet_post_write(ws_f, last_row)
    _add_rilasciato_validation(ws_f, last_row)

    ws_c.cell(7, 16, "\xa0")
    ws_c.cell(8, 16, "\xa0")
    ws_f.cell(7, 16, "\xa0")
    ws_f.cell(8, 16, "\xa0")

    wb.save(OUT_FABBISOGNO)
    wb.close()
    return OUT_FABBISOGNO


def generate_dataiku() -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if TEMPLATE_DATAIKU.exists():
        shutil.copy2(TEMPLATE_DATAIKU, OUT_DATAIKU)
    elif not OUT_DATAIKU.exists():
        raise FileNotFoundError(
            f"Template not found: {TEMPLATE_DATAIKU} and no existing {OUT_DATAIKU}"
        )

    wb = load_workbook(OUT_DATAIKU)
    ws = wb.active

    ws["A1"] = "DATAIKU RSC - COLLAUDO / PRODUCTION (SIV - iTree)"

    # Tabella1: solo iTree + riga TOT. (A2:E4)
    ws.cell(3, 1, "iTree")
    ws.cell(3, 2, 1)
    ws.cell(3, 3, 2)
    ws.cell(3, 4, 4)
    ws.cell(3, 5, 8)

    for r in range(5, 22):
        for c in range(1, 6):
            ws.cell(r, c).value = None

    ws.cell(4, 1).value = "TOT."
    ws.cell(4, 2).value = "=SUM(Tabella1[CPU Request (vCore)])"
    ws.cell(4, 3).value = "=SUM(Tabella1[CPU Limit (vCore)])"
    ws.cell(4, 4).value = "=SUM(Tabella1[RAM Request (GB)])"
    ws.cell(4, 5).value = "=SUM(Tabella1[RAM Limit (GB)])"

    if "Tabella1" in ws.tables:
        ws.tables["Tabella1"].ref = "A2:E4"

    # Tabella24: solo iTree — Light workload, batch per comune (G2:N3)
    ws.cell(3, 7, "Light")
    ws.cell(3, 8, "SIV")
    ws.cell(3, 9, "iTree — servizi ecosistemici (CO2, stormwater, energy)")
    ws.cell(3, 10, 2)
    ws.cell(3, 11, 8)
    ws.cell(3, 12, 2)
    ws.cell(3, 13, 4)
    ws.cell(3, 14, 16)

    for r in range(4, 6):
        for c in range(7, 15):
            ws.cell(r, c).value = None

    if "Tabella24" in ws.tables:
        ws.tables["Tabella24"].ref = "G2:N3"

    ws["A22"] = (
        "Note: Dimensionamento Dataiku per modello iTree (ecosystem services). "
        "Batch per comune su inventario arboreo censito (~8.4M alberi nazionali)."
    )

    ws["B25"] = 4
    ws["B26"] = 16
    ws["E25"] = 8
    ws["E26"] = 32

    ws["A29"] = (
        "Quota iTree condivisa ambiente Dataiku SIV. "
        "Elaborazioni batch leggero/medio; da rivedere con volumetrie effettive per comune."
    )

    wb.save(OUT_DATAIKU)
    wb.close()
    return OUT_DATAIKU


def main() -> None:
    fab = generate_fabbisogno()
    dku = generate_dataiku()

    shutil.copy2(fab, DESKTOP / fab.name)
    shutil.copy2(dku, DESKTOP / dku.name)

    print(f"Generated (template clone): {fab}")
    print(f"Generated (template clone): {dku}")
    print(f"Copied to Desktop: {fab.name}, {dku.name}")


if __name__ == "__main__":
    main()
